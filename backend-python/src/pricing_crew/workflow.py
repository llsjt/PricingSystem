from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, func
from sqlalchemy.orm import Session

from pricing_crew.config.runtime import settings
from pricing_crew.db import init_database, seed_realistic_demo_data
from pricing_crew.db.database import SessionLocal
from pricing_crew.db.models import BizProduct, BizSalesDaily, DecAgentLog, DecResult, DecTask
from pricing_crew.agent_logic import (
    run_data_analysis_agent,
    run_manager_coordinator_agent,
    run_market_intel_agent,
    run_risk_control_agent,
)
from pricing_crew.schemas import AnalysisRequest, CompetitorData, ProductBase, RiskData, SalesData
from pricing_crew.ws_manager import manager

logger = logging.getLogger(__name__)


def _f(value: Any, default: float = 0.0) -> float:
    """把 Decimal、None 等值统一转成 float，减少金额和销量字段的类型分支。"""
    if isinstance(value, Decimal):
        return float(value)
    if value is None:
        return default
    return float(value)


class WorkflowService:
    DATA_ROLE = "数据分析"
    MARKET_ROLE = "市场情报"
    RISK_ROLE = "风险控制"
    MANAGER_ROLE = "决策经理"

    def bootstrap(self) -> None:
        """启动服务时初始化数据库，并按配置写入演示数据。"""
        init_database()
        if settings.bootstrap_demo_data:
            seed_realistic_demo_data()

    def _next_id(self, db: Session, model) -> int:
        """在演示环境中通过“当前最大值 + 1”生成主键。"""
        return int(db.query(func.max(model.id)).scalar() or 0) + 1

    def _dec(self, value: Any, default: float = 0.0) -> Decimal:
        """把输入安全地转成 Decimal，统一数据库金额字段写入格式。"""
        try:
            return Decimal(str(float(value if value is not None else default)))
        except Exception:
            return Decimal(str(default))

    def _parse_time(self, text: str) -> Optional[datetime]:
        """兼容完整时间和纯日期两种输入格式。"""
        s = (text or "").strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        return None

    def _fill_sales(self, db: Session, product_id: int, price: float, monthly_sales: int, days: int = 30) -> None:
        """为商品补齐演示销量明细，确保趋势页和定价分析有基础数据可用。"""
        db.query(BizSalesDaily).filter(BizSalesDaily.product_id == product_id).delete(synchronize_session=False)
        baseline = max(monthly_sales / 30.0, 1.0)
        next_id = self._next_id(db, BizSalesDaily)
        today = date.today()
        for i in range(days):
            d = today - timedelta(days=days - i - 1)
            factor = 1.15 if d.weekday() in {4, 5} else 0.95 if d.weekday() == 0 else 1.0
            sold = max(1, int(round(baseline * factor)))
            db.add(
                BizSalesDaily(
                    id=next_id,
                    product_id=product_id,
                    stat_date=d,
                    daily_sales=sold,
                    daily_revenue=self._dec(sold * price),
                )
            )
            next_id += 1

    def get_product_list(self, page: int, size: int, keyword: str = "", data_source: str = "") -> Dict[str, Any]:
        """查询商品列表，并按前端字段格式返回分页结果。"""
        db: Session = SessionLocal()
        try:
            q = db.query(BizProduct)
            if keyword:
                q = q.filter(BizProduct.title.ilike(f"%{keyword}%"))
            if data_source:
                q = q.filter(BizProduct.source == data_source)
            total = int(q.count() or 0)
            rows = q.order_by(BizProduct.updated_at.desc(), BizProduct.id.desc()).offset((page - 1) * size).limit(size).all()
            return {
                "records": [
                    {
                        "id": int(r.id),
                        "title": r.title,
                        "category": r.category,
                        "costPrice": _f(r.cost_price),
                        "marketPrice": _f(r.market_price) if r.market_price is not None else None,
                        "currentPrice": _f(r.current_price),
                        "stock": int(r.stock or 0),
                        "source": r.source,
                        "monthlySales": int(r.monthly_sales or 0),
                        "clickRate": _f(r.click_rate),
                        "conversionRate": _f(r.conversion_rate),
                        "updatedAt": r.updated_at.isoformat() if r.updated_at else None,
                    }
                    for r in rows
                ],
                "total": total,
            }
        finally:
            db.close()

    def add_product_manual(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """处理手工新增商品，并同步生成一段基础销量数据。"""
        db: Session = SessionLocal()
        try:
            title = str(payload.get("title") or "").strip()
            if not title:
                raise ValueError("商品标题不能为空")
            p = BizProduct(
                id=self._next_id(db, BizProduct),
                title=title,
                category=str(payload.get("category") or "").strip() or None,
                cost_price=self._dec(payload.get("costPrice")),
                market_price=self._dec(payload.get("marketPrice")) if payload.get("marketPrice") is not None else None,
                current_price=self._dec(payload.get("currentPrice")),
                stock=int(payload.get("stock") or 0),
                monthly_sales=int(payload.get("monthlySales") or 0),
                click_rate=self._dec(payload.get("clickRate")),
                conversion_rate=self._dec(payload.get("conversionRate")),
                source=str(payload.get("source") or "MANUAL"),
                updated_at=datetime.utcnow(),
            )
            db.add(p)
            db.flush()
            self._fill_sales(db, int(p.id), _f(p.current_price), int(p.monthly_sales or 0), 30)
            db.commit()
            return {"id": int(p.id), "title": p.title}
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

    def batch_delete_products(self, ids: List[int]) -> int:
        """批量删除商品及其销量明细，避免留下孤立趋势数据。"""
        db: Session = SessionLocal()
        try:
            id_set = {int(x) for x in ids if str(x).strip()}
            if not id_set:
                return 0
            n = db.query(BizProduct).filter(BizProduct.id.in_(id_set)).delete(synchronize_session=False)
            db.query(BizSalesDaily).filter(BizSalesDaily.product_id.in_(id_set)).delete(synchronize_session=False)
            db.commit()
            return int(n or 0)
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

    def build_product_template(self) -> bytes:
        """生成商品导入模板，供前端下载。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "products"
        ws.append(["title", "category", "costPrice", "marketPrice", "currentPrice", "stock", "monthlySales", "clickRate", "conversionRate", "source"])
        ws.append(["示例商品", "数码配件", 59.0, 99.0, 79.0, 100, 300, 0.07, 0.03, "IMPORT"])
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def import_products_from_excel(self, data: bytes, filename: str = "") -> Dict[str, Any]:
        """导入 Excel 商品数据，并为每个新商品补齐演示销量明细。"""
        if not data:
            raise ValueError("上传文件为空")
        if filename.lower().endswith(".xls"):
            raise ValueError("暂不支持 .xls，请改用 .xlsx")
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        if not rows:
            raise ValueError("导入文件没有有效内容")
        head = [str(x or "").strip() for x in rows[0]]
        idx = {k: i for i, k in enumerate(head)}
        for col in ("title", "costPrice", "currentPrice"):
            if col not in idx:
                raise ValueError(f"缺少列: {col}")
        ok, skip = 0, 0
        db: Session = SessionLocal()
        try:
            for row in rows[1:]:
                title = str(row[idx["title"]] or "").strip()
                if not title:
                    skip += 1
                    continue
                p = BizProduct(
                    id=self._next_id(db, BizProduct),
                    title=title,
                    category=str(row[idx.get("category", -1)] or "").strip() if "category" in idx else None,
                    cost_price=self._dec(row[idx["costPrice"]]),
                    market_price=self._dec(row[idx["marketPrice"]]) if "marketPrice" in idx and row[idx["marketPrice"]] is not None else None,
                    current_price=self._dec(row[idx["currentPrice"]]),
                    stock=int(row[idx["stock"]] or 0) if "stock" in idx else 0,
                    monthly_sales=int(row[idx["monthlySales"]] or 0) if "monthlySales" in idx else 0,
                    click_rate=self._dec(row[idx["clickRate"]]) if "clickRate" in idx else Decimal("0.0"),
                    conversion_rate=self._dec(row[idx["conversionRate"]]) if "conversionRate" in idx else Decimal("0.0"),
                    source=str(row[idx["source"]] or "IMPORT") if "source" in idx else "IMPORT",
                    updated_at=datetime.utcnow(),
                )
                db.add(p)
                db.flush()
                self._fill_sales(db, int(p.id), _f(p.current_price), int(p.monthly_sales or 0), 30)
                ok += 1
            db.commit()
        except Exception:
            db.rollback()
            raise
        finally:
            wb.close()
            db.close()
        return {"imported": ok, "skipped": skip, "message": f"导入完成：成功 {ok} 条，跳过 {skip} 条"}

    def get_product_trend(self, product_id: int, days: int = 30) -> Dict[str, Any]:
        """返回商品趋势图和趋势卡片需要的销售、利润汇总数据。"""
        days = max(7, min(int(days or 30), 180))
        db: Session = SessionLocal()
        try:
            p = db.query(BizProduct).filter(BizProduct.id == product_id).first()
            if p is None:
                raise ValueError(f"商品 {product_id} 不存在")
            end_d = date.today()
            start_d = end_d - timedelta(days=days - 1)
            rows = db.query(BizSalesDaily).filter(BizSalesDaily.product_id == product_id, BizSalesDaily.stat_date >= start_d, BizSalesDaily.stat_date <= end_d).order_by(BizSalesDaily.stat_date.asc()).all()
            if not rows:
                self._fill_sales(db, product_id, _f(p.current_price), int(p.monthly_sales or 0), days)
                db.commit()
                rows = db.query(BizSalesDaily).filter(BizSalesDaily.product_id == product_id, BizSalesDaily.stat_date >= start_d, BizSalesDaily.stat_date <= end_d).order_by(BizSalesDaily.stat_date.asc()).all()
            sales = [int(r.daily_sales or 0) for r in rows]
            dates = [r.stat_date.strftime("%m-%d") for r in rows]
            visitors = [max(int(round(s / max(_f(p.conversion_rate, 0.04), 0.01))), s) for s in sales]
            conv = [round(s / max(visitors[i], 1), 4) for i, s in enumerate(sales)]
            w1, w0 = sales[-7:] or [0], (sales[-14:-7] or [0])
            g = sum(w1) / max(len(w1), 1) - sum(w0) / max(len(w0), 1)
            gr = g / max(sum(w0) / max(len(w0), 1), 1.0)
            m1, m0 = sum(sales[-30:]), sum(sales[-60:-30]) if len(sales) > 30 else max(sum(sales), 1)
            mg, mgr = m1 - m0, (m1 - m0) / max(m0, 1.0)
            up = _f(p.current_price) - _f(p.cost_price)
            return {"dates": dates, "sales": sales, "visitors": visitors, "conversionRates": conv, "dailySalesGrowth": round(g, 2), "dailySalesGrowthRate": round(gr, 4), "monthlySalesGrowth": round(mg, 2), "monthlySalesGrowthRate": round(mgr, 4), "dailyProfitGrowth": round(g * up, 2), "dailyProfitGrowthRate": round(gr, 4), "monthlyProfitGrowth": round(mg * up, 2), "monthlyProfitGrowthRate": round(mgr, 4)}
        finally:
            db.close()

    def task_exists(self, task_id: int) -> bool:
        """判断指定任务是否已经存在，避免重复创建同一个任务。"""
        db: Session = SessionLocal()
        try:
            return bool(db.query(DecTask).filter(DecTask.id == task_id).first())
        finally:
            db.close()

    def create_task(
        self,
        product_ids: List[int],
        strategy_goal: str,
        constraints: str,
        predefined_task_id: Optional[int] = None,
    ) -> int:
        """创建定价任务主记录，供 Python 与 Java 共同引用。"""
        db: Session = SessionLocal()
        try:
            if predefined_task_id is not None:
                exists = db.query(DecTask).filter(DecTask.id == predefined_task_id).first()
                if exists:
                    return int(exists.id)
            products = db.query(BizProduct).filter(BizProduct.id.in_(product_ids)).all()
            task = DecTask(
                id=predefined_task_id if predefined_task_id is not None else self._next_id(db, DecTask),
                task_no=f"TASK-{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}",
                product_names="、".join(p.title for p in products),
                strategy_type=strategy_goal,
                constraints=constraints,
                status="PENDING",
                product_count=len(product_ids),
            )
            db.add(task)
            db.commit()
            db.refresh(task)
            return int(task.id)
        finally:
            db.close()

    def update_task_status(self, task_id: int, status: str) -> None:
        """更新任务状态，并在关键状态下补齐开始/完成时间。"""
        db: Session = SessionLocal()
        try:
            t = db.query(DecTask).filter(DecTask.id == task_id).first()
            if not t:
                return
            t.status = status
            if status == "RUNNING" and not t.started_at:
                t.started_at = datetime.utcnow()
            if status in {"COMPLETED", "FAILED"}:
                t.completed_at = datetime.utcnow()
            db.commit()
        except Exception:
            db.rollback()
        finally:
            db.close()

    def parse_constraint_bundle(self, constraints: str) -> Dict[str, Any]:
        """从自然语言约束里提取利润率、折扣、价格上下限等结构化参数。"""
        parsed: Dict[str, Any] = {"raw_text": constraints or "", "min_profit_margin": None, "min_profit_markup": None, "max_discount_allowed": None, "price_floor": None, "price_ceiling": None, "summary": []}
        if not constraints:
            return parsed
        m = re.search(r"(?:利润率|毛利率|profit\\s*margin)\\D*(\\d+(?:\\.\\d+)?)\\s*%", constraints, re.I)
        if m:
            parsed["min_profit_margin"] = float(m.group(1)) / 100.0
            parsed["summary"].append(f"最低利润率 >= {m.group(1)}%")
        m = re.search(r"(?:加价率|加成|markup)\\D*(\\d+(?:\\.\\d+)?)\\s*%", constraints, re.I)
        if m:
            parsed["min_profit_markup"] = float(m.group(1)) / 100.0
        m = re.search(r"(?:降价幅度|折扣|优惠|discount)\\D*(\\d+(?:\\.\\d+)?)\\s*%", constraints, re.I)
        if m:
            parsed["max_discount_allowed"] = max(0.0, 1 - float(m.group(1)) / 100.0)
        m = re.search(r"(?:最低售价|最低价格|price\\s*floor)\\D*(\\d+(?:\\.\\d+)?)", constraints, re.I)
        if m:
            parsed["price_floor"] = float(m.group(1))
        m = re.search(r"(?:最高售价|最高价格|price\\s*ceiling)\\D*(\\d+(?:\\.\\d+)?)", constraints, re.I)
        if m:
            parsed["price_ceiling"] = float(m.group(1))
        return parsed

    def build_risk_data(self, product: BizProduct, parsed: Dict[str, Any]) -> RiskData:
        """把数据库商品信息和解析后的约束，拼成风控智能体需要的输入对象。"""
        return RiskData(min_profit_margin=parsed.get("min_profit_margin") or (0.2 if (product.category or "") == "数码配件" else 0.18), min_profit_markup=parsed.get("min_profit_markup"), target_profit_margin=0.30, max_discount_allowed=parsed.get("max_discount_allowed"), price_floor=parsed.get("price_floor"), price_ceiling=parsed.get("price_ceiling"), enforce_hard_constraints=True, constraint_summary=list(parsed.get("summary") or []))

    def _to_cn_action(self, action: str) -> str:
        mapping = {
            "maintain": "维持原价",
            "discount": "小幅降价",
            "increase": "适度提价",
            "clearance": "清仓促销",
            "premium": "上探价格带",
            "penetrate": "渗透降价",
            "price_war": "竞争性降价",
        }
        return mapping.get(str(action or "").lower(), "维持原价")

    def _to_cn_trend(self, trend: str) -> str:
        mapping = {"rising": "上升", "declining": "下降", "stable": "平稳"}
        return mapping.get(str(trend or "").lower(), "平稳")

    def _to_cn_inventory(self, status: str) -> str:
        mapping = {
            "severe_overstock": "严重积压",
            "overstock": "库存偏高",
            "normal": "库存正常",
            "tight": "库存偏紧",
        }
        return mapping.get(str(status or "").lower(), "库存正常")

    def _to_cn_competition(self, level: str) -> str:
        mapping = {"fierce": "激烈", "moderate": "中等", "low": "较低"}
        return mapping.get(str(level or "").lower(), "中等")

    def _to_cn_price_position(self, position: str) -> str:
        mapping = {"premium": "偏高价位", "mid-range": "中位价带", "budget": "偏低价位"}
        return mapping.get(str(position or "").lower(), "中位价带")

    def _to_cn_risk_level(self, level: str) -> str:
        mapping = {"high": "高风险", "medium": "中风险", "low": "低风险"}
        return mapping.get(str(level or "").lower(), "中风险")

    def _format_trend_change(self, score: float) -> str:
        return f"近7日相对近30日的销量变化率为{float(score):.1%}"

    def _format_discount_guidance(self, min_rate: float, max_rate: float) -> str:
        min_rate = float(min_rate)
        max_rate = float(max_rate)
        if abs(min_rate - 1.0) < 1e-6 and abs(max_rate - 1.0) < 1e-6:
            return "当前建议维持原价，不打折。"
        if min_rate <= 1.0 and max_rate <= 1.0:
            min_drop = max(0.0, 1.0 - max_rate)
            max_drop = max(0.0, 1.0 - min_rate)
            return f"建议折扣系数区间为{min_rate:.2f}-{max_rate:.2f}（约降价{min_drop:.0%}-{max_drop:.0%}）。"
        return f"建议价格调整系数区间为{min_rate:.2f}-{max_rate:.2f}。"

    def _format_live_market_status(self, available: bool) -> str:
        if available:
            return "实时竞品数据可用，样本量满足分析要求。"
        return "实时竞品数据暂不可用或样本不足，结果已按保守策略处理。"

    def _humanize_agent_text(self, text: str) -> str:
        """把内部术语、英文标签和技术字段替换成前端用户能读懂的中文文案。"""
        normalized = str(text or "")
        replacements = {
            "把握程度": "可信度",
            "置信度": "可信度",
            "Agent": "智能体",
            "agent": "智能体",
            "simulated_market_dataset": "模拟市场样本库",
            "MAX_PROFIT": "利润优先",
            "CLEARANCE": "清仓促销",
            "MARKET_SHARE": "市场份额优先",
            "maintain": "维持原价",
            "discount": "小幅降价",
            "increase": "适度提价",
            "premium": "上探价格带",
            "penetrate": "渗透定价",
            "price_war": "竞争性降价",
            "rising": "上升",
            "declining": "下降",
            "stable": "平稳",
            "fierce": "激烈",
            "moderate": "中等",
            "high": "高",
            "medium": "中",
            "low": "低",
            "True": "是",
            "False": "否",
            "true": "是",
            "false": "否",
            "required_min_price": "最低安全价",
            "current_margin": "当前毛利率",
            "risk_score": "风险指数",
            "constraint_conflict": "约束冲突",
            "manual_review": "人工复核",
            "risk_boundary": "风险边界",
            "forced_adjustment": "强制调价",
            "three_agent_proposals": "三方提案",
            "manager_coordination": "经理协调",
            "profit_validation": "利润校验",
            "竞争评分": "竞争热度",
            "风险分": "风险指数",
        }
        for old, new in replacements.items():
            normalized = normalized.replace(old, new)
        # 前端展示文案统一去掉等号，避免机器化表达
        normalized = normalized.replace("=", "：")
        return normalized

    def compose_data_agent_output(self, name: str, result) -> str:
        """把数据分析结果拼成前端流式展示用的话术。"""
        reasons = "；".join(result.recommendation_reasons[:2]) if result.recommendation_reasons else "暂无额外补充依据。"
        discount_text = self._format_discount_guidance(
            result.recommended_discount_range[0],
            result.recommended_discount_range[1],
        )
        if str(result.recommended_action or "").lower() == "maintain" and "维持原价" in discount_text:
            conclusion_text = f"简要结论：{discount_text}可信度 {result.confidence:.0%}。"
        else:
            conclusion_text = (
                f"简要结论：建议{self._to_cn_action(result.recommended_action)}，"
                f"{discount_text}可信度 {result.confidence:.0%}。"
            )
        return (
            f"思考过程：{result.thinking_process or '已完成经营数据读取并完成趋势判断。'}\n"
            "依据：\n"
            f"1. {result.reasoning}\n"
            f"2. 销量趋势为{self._to_cn_trend(result.sales_trend)}，{self._format_trend_change(result.sales_trend_score)}；"
            f"库存状态为{self._to_cn_inventory(result.inventory_status)}，库存平衡度为{result.inventory_health_score:.1f}分。\n"
            f"3. {reasons}\n"
            f"{conclusion_text}"
        )

    def compose_market_agent_output(self, name: str, result) -> str:
        """把市场情报结果拼成前端流式展示用的话术。"""
        reasons = "；".join(result.suggestion_reasons[:2]) if result.suggestion_reasons else "暂无额外补充依据。"
        data_sources = "、".join(result.data_sources[:2]) if result.data_sources else "无可用市场源"
        live_market_text = self._format_live_market_status(bool(result.analysis_details.get("live_market_available")))
        return (
            f"思考过程：{result.thinking_process or '已完成竞品抓取与市场对位分析。'}\n"
            "依据：\n"
            f"1. {result.reasoning}\n"
            f"2. 竞争强度为{self._to_cn_competition(result.competition_level)}，竞争热度为{result.competition_score:.2f}分；"
            f"价格位置为{self._to_cn_price_position(result.price_position)}，价格分位为{result.price_percentile:.1%}。\n"
            f"3. 数据来源：{data_sources}；{live_market_text}；{reasons}\n"
            f"简要结论：建议{self._to_cn_action(result.market_suggestion)}，可信度 {result.confidence:.0%}。"
        )

    def compose_risk_agent_output(self, name: str, result) -> str:
        """把风控结果拼成前端流式展示用的话术。"""
        reasons = "；".join(result.recommendation_reasons[:2]) if result.recommendation_reasons else "暂无额外补充依据。"
        warning_text = "；".join(result.warnings[:2]) if result.warnings else "暂无阻断型风险预警。"
        return (
            f"思考过程：{result.thinking_process or '已完成利润底线与约束合规校验。'}\n"
            "依据：\n"
            f"1. {result.reasoning}\n"
            f"2. 最低安全价为{result.required_min_price:.2f}元，当前毛利率为{result.current_profit_margin:.2%}，"
            f"风险等级为{self._to_cn_risk_level(result.risk_level)}，风险指数为{result.risk_score:.1f}分。\n"
            f"3. {reasons}；{warning_text}\n"
            f"简要结论：建议{self._to_cn_action(result.recommendation)}，允许最大折扣系数 {result.max_safe_discount:.2f}。"
        )

    def compose_manager_agent_output(self, name: str, result) -> str:
        """把经理协调结果拼成前端流式展示用的话术。"""
        factors = "、".join(result.key_factors[:3]) if result.key_factors else "暂无关键因子"
        warnings = "；".join(result.warnings[:2]) if result.warnings else "无"
        next_step = result.execution_plan[0].action if result.execution_plan else "按建议价格上线并持续监测。"
        return (
            f"思考过程：{result.thinking_process or '已汇总三位智能体的分析并做约束校验。'}\n"
            "依据：\n"
            f"1. {result.reasoning}\n"
            f"2. 核心因子：{factors}；风险提示：{warnings}。\n"
            f"3. 执行动作：{next_step}\n"
            f"简要结论：最终建议{self._to_cn_action(result.decision)}，建议价格 {result.suggested_price:.2f}元，可信度 {result.confidence:.0%}。"
        )

    async def execute_task(self, task_id: int, product_ids: List[int], strategy_goal: str, constraints: str) -> None:
        """按商品逐个执行完整的四智能体定价流程。"""
        self.update_task_status(task_id, "RUNNING")
        for product_id in product_ids:
            await self.process_single_product(task_id, product_id, strategy_goal, constraints)
        self.update_task_status(task_id, "COMPLETED")
        await manager.broadcast(json.dumps({"type": "complete", "task_id": task_id}, ensure_ascii=False), str(task_id))

    async def process_single_product(self, task_id: int, product_id: int, strategy_goal: str, constraints: str) -> None:
        """处理单个商品的完整定价链路：取数、跑四个智能体、落库并推送。"""
        db: Session = SessionLocal()
        try:
            # 仅在执行 agent 时从数据库读取商品上下文，避免在 Python 层承载通用业务逻辑。
            # 先从数据库读取当前商品快照，并据此组装四个智能体共享的分析请求。
            product = db.query(BizProduct).filter(BizProduct.id == product_id).first()
            if not product:
                raise ValueError(f"商品 {product_id} 不存在")
            parsed = self.parse_constraint_bundle(constraints)
            req = AnalysisRequest(task_id=str(task_id), product=ProductBase(product_id=str(product.id), product_name=product.title, category=product.category or "未分类", current_price=_f(product.current_price), cost=_f(product.cost_price), original_price=_f(product.market_price) if product.market_price is not None else None, stock=int(product.stock or 0), stock_age_days=max(14, int((int(product.stock or 0) / max(int(product.monthly_sales or 1) / 30.0, 1.0)) * 1.25))), sales_data=SalesData(), competitor_data=CompetitorData(competitors=[]), risk_data=self.build_risk_data(product, parsed), customer_reviews=[], strategy_goal=strategy_goal, strategy_constraints=constraints, business_context={"product_id": int(product.id), "prefer_db_tools": True, "parsed_constraints": parsed})

            # 每个 agent 完成后立即推流，避免前端长时间无输出。
            # 四个智能体按照“数据 -> 市场 -> 风控 -> 经理”的固定顺序依次发言。
            await self.stream_thought(task_id, self.DATA_ROLE, 1, "", product_id, emit_end=False)
            data_result = await asyncio.to_thread(run_data_analysis_agent, req)
            data_text = self._humanize_agent_text(self.compose_data_agent_output(product.title, data_result))
            await self.stream_thought(task_id, self.DATA_ROLE, 1, data_text, product_id, emit_start=False)
            self._save_log(db, task_id, product_id, self.DATA_ROLE, 1, data_text)

            await self.stream_thought(task_id, self.MARKET_ROLE, 2, "", product_id, emit_end=False)
            market_result = await asyncio.to_thread(run_market_intel_agent, req)
            if not bool(market_result.analysis_details.get("live_market_available")):
                failure_reasons = market_result.analysis_details.get("failure_reasons") or market_result.limitations or []
                failure_items = [str(item) for item in failure_reasons if item]
                reason_text = "；".join(failure_items[:2]) if failure_items else "未知原因"
                await manager.broadcast(
                    json.dumps(
                        {
                            "type": "market_unavailable",
                            "task_id": task_id,
                            "product_id": product_id,
                            "message": f"商品 {product_id} 无法生成竞品市场样本：{reason_text}",
                        },
                        ensure_ascii=False,
                    ),
                    str(task_id),
                )
            market_text = self._humanize_agent_text(self.compose_market_agent_output(product.title, market_result))
            await self.stream_thought(task_id, self.MARKET_ROLE, 2, market_text, product_id, emit_start=False)
            self._save_log(db, task_id, product_id, self.MARKET_ROLE, 2, market_text)

            await self.stream_thought(task_id, self.RISK_ROLE, 3, "", product_id, emit_end=False)
            risk_result = await asyncio.to_thread(run_risk_control_agent, req)
            risk_text = self._humanize_agent_text(self.compose_risk_agent_output(product.title, risk_result))
            await self.stream_thought(task_id, self.RISK_ROLE, 3, risk_text, product_id, emit_start=False)
            self._save_log(db, task_id, product_id, self.RISK_ROLE, 3, risk_text)

            await self.stream_thought(task_id, self.MANAGER_ROLE, 4, "", product_id, emit_end=False)
            final_result = await asyncio.to_thread(run_manager_coordinator_agent, req, data_result, market_result, risk_result)
            manager_text = self._humanize_agent_text(self.compose_manager_agent_output(product.title, final_result))
            await self.stream_thought(task_id, self.MANAGER_ROLE, 4, manager_text, product_id, emit_start=False)
            self._save_log(db, task_id, product_id, self.MANAGER_ROLE, 4, manager_text)

            self._save_result(db, task_id, product_id, final_result)
            await manager.broadcast(json.dumps({"type": "result", "task_id": task_id, "product_id": product_id, "data": final_result.model_dump(mode="json")}, ensure_ascii=False), str(task_id))
        finally:
            db.close()

    def _save_log(self, db: Session, task_id: int, product_id: int, role: str, order: int, content: str) -> None:
        """保存单个智能体的发言日志，供前端任务详情页回放。"""
        text = f"商品 {product_id}\n{content}"
        db.add(
            DecAgentLog(
                id=self._next_id(db, DecAgentLog),
                task_id=task_id,
                role_name=role,
                speak_order=order,
                thought_content=text,
            )
        )
        db.commit()

    def _save_result(self, db: Session, task_id: int, product_id: int, final) -> None:
        """保存最终定价结果，若已有记录则执行覆盖更新。"""
        row = db.query(DecResult).filter(DecResult.task_id == task_id, DecResult.product_id == product_id).first()
        if row:
            row.decision, row.discount_rate, row.suggested_price, row.profit_change, row.core_reasons = final.decision, self._dec(final.discount_rate), self._dec(final.suggested_price), self._dec(final.expected_outcomes.profit_change if final.expected_outcomes else 0), final.core_reasons
        else:
            db.add(DecResult(id=self._next_id(db, DecResult), task_id=task_id, product_id=product_id, decision=final.decision, discount_rate=self._dec(final.discount_rate), suggested_price=self._dec(final.suggested_price), profit_change=self._dec(final.expected_outcomes.profit_change if final.expected_outcomes else 0), core_reasons=final.core_reasons, is_accepted=False, adopt_status="PENDING"))
        db.commit()

    async def stream_thought(
        self,
        task_id: int,
        role: str,
        step: int,
        content: str,
        product_id: int,
        emit_start: bool = True,
        emit_end: bool = True,
    ) -> None:
        """把智能体输出拆成小片段逐步推送，模拟实时协作讨论过程。"""
        if emit_start:
            await manager.broadcast(
                json.dumps(
                    {
                        "is_stream": True,
                        "is_start": True,
                        "agent_role": role,
                        "step_order": step,
                        "product_id": product_id,
                        "timestamp": datetime.now().strftime("%H:%M:%S"),
                    },
                    ensure_ascii=False,
                ),
                str(task_id),
            )

        chunk_size = max(4, int(settings.websocket_chunk_size or 24))
        raw_delay = max(0.0, float(settings.websocket_chunk_delay or 0.0))
        min_stream_seconds = max(0.0, float(settings.websocket_min_stream_seconds or 0.0))
        chunk_count = max(1, (len(content) + chunk_size - 1) // chunk_size)
        # 让每条 agent 消息有可感知的流式时长，避免“瞬间刷完”的体感。
        chunk_delay = max(raw_delay, 0.03, (min_stream_seconds / chunk_count))

        for i in range(0, len(content), chunk_size):
            await manager.broadcast(
                json.dumps(
                    {
                        "is_stream": True,
                        "agent_role": role,
                        "step_order": step,
                        "product_id": product_id,
                        "thought_content": content[i : i + chunk_size],
                        "timestamp": datetime.now().strftime("%H:%M:%S"),
                    },
                    ensure_ascii=False,
                ),
                str(task_id),
            )
            await asyncio.sleep(chunk_delay)

        if emit_end:
            await manager.broadcast(
                json.dumps(
                    {
                        "is_stream": True,
                        "is_end": True,
                        "agent_role": role,
                        "step_order": step,
                        "product_id": product_id,
                        "timestamp": datetime.now().strftime("%H:%M:%S"),
                    },
                    ensure_ascii=False,
                ),
                str(task_id),
            )

    def get_task_results(self, task_id: int) -> List[Dict[str, Any]]:
        """读取任务结果列表，并补齐原价、采纳状态等前端字段。"""
        db: Session = SessionLocal()
        try:
            rows = db.query(DecResult, BizProduct).join(BizProduct, BizProduct.id == DecResult.product_id).filter(DecResult.task_id == task_id).order_by(DecResult.id.asc()).all()
            results: List[Dict[str, Any]] = []
            for r, p in rows:
                suggested_price = _f(r.suggested_price)
                discount_rate = _f(r.discount_rate)
                current_price = _f(p.current_price)
                original_price = round(suggested_price / discount_rate, 2) if discount_rate > 0 else round(current_price, 2)
                results.append(
                    {
                        "id": int(r.id),
                        "taskId": int(r.task_id),
                        "productId": int(r.product_id),
                        "productTitle": p.title,
                        "decision": r.decision,
                        "originalPrice": original_price,
                        "suggestedPrice": suggested_price,
                        "profitChange": _f(r.profit_change),
                        "discountRate": discount_rate,
                        "coreReasons": r.core_reasons,
                        "isAccepted": bool(r.is_accepted),
                        "adoptStatus": r.adopt_status or "PENDING",
                        "createdAt": r.created_at.isoformat() if r.created_at else None,
                    }
                )
            return results
        finally:
            db.close()

    def get_task_logs(self, task_id: int) -> List[Dict[str, Any]]:
        db: Session = SessionLocal()
        try:
            rows = db.query(DecAgentLog).filter(DecAgentLog.task_id == task_id).order_by(DecAgentLog.speak_order.asc(), DecAgentLog.id.asc()).all()
            return [{"id": int(x.id), "roleName": x.role_name, "thoughtContent": x.thought_content or "", "speakOrder": int(x.speak_order or 0), "createdAt": x.created_at.isoformat() if x.created_at else None} for x in rows]
        finally:
            db.close()

    def get_task_archive(self, page: int, size: int, status: str = "", strategy_type: str = "", start_time: str = "", end_time: str = "", sort_order: str = "desc") -> Dict[str, Any]:
        db: Session = SessionLocal()
        try:
            q = db.query(DecTask)
            fs = []
            if status:
                fs.append(DecTask.status == status)
            if strategy_type:
                fs.append(DecTask.strategy_type == strategy_type)
            st, et = self._parse_time(start_time), self._parse_time(end_time)
            if st:
                fs.append(DecTask.created_at >= st)
            if et:
                fs.append(DecTask.created_at <= et)
            if fs:
                q = q.filter(and_(*fs))
            total = int(q.count() or 0)
            order = DecTask.created_at.asc() if str(sort_order).lower() == "asc" else DecTask.created_at.desc()
            rows = q.order_by(order, DecTask.id.desc()).offset((page - 1) * size).limit(size).all()
            return {"records": [{"id": int(t.id), "taskNo": t.task_no, "productNames": t.product_names or "", "strategyType": t.strategy_type, "status": t.status, "createdAt": t.created_at.isoformat() if t.created_at else None} for t in rows], "total": total}
        finally:
            db.close()

    def get_task_stats(self, strategy_type: str = "", start_time: str = "", end_time: str = "") -> Dict[str, int]:
        rows = self.get_task_archive(1, 100000, "", strategy_type, start_time, end_time, "desc")["records"]
        return {"completed": sum(1 for x in rows if x["status"] == "COMPLETED"), "running": sum(1 for x in rows if x["status"] == "RUNNING"), "failed": sum(1 for x in rows if x["status"] == "FAILED"), "pending": sum(1 for x in rows if x["status"] == "PENDING")}

    def get_task_comparison(self, task_id: int) -> List[Dict[str, Any]]:
        db: Session = SessionLocal()
        try:
            rows = db.query(DecResult, BizProduct).join(BizProduct, BizProduct.id == DecResult.product_id).filter(DecResult.task_id == task_id).order_by(DecResult.id.asc()).all()
            return [{"resultId": int(r.id), "taskId": int(r.task_id), "productId": int(r.product_id), "productTitle": p.title, "originalPrice": _f(p.current_price), "suggestedPrice": _f(r.suggested_price), "profitChange": _f(r.profit_change), "discountRate": _f(r.discount_rate), "adoptStatus": r.adopt_status or ("ADOPTED" if r.is_accepted else "PENDING"), "rejectReason": r.reject_reason, "createdAt": r.created_at.isoformat() if r.created_at else None} for r, p in rows]
        finally:
            db.close()

    def apply_result(self, result_id: int) -> Dict[str, Any]:
        db: Session = SessionLocal()
        try:
            r = db.query(DecResult).filter(DecResult.id == result_id).first()
            if not r:
                raise ValueError(f"结果 {result_id} 不存在")
            p = db.query(BizProduct).filter(BizProduct.id == r.product_id).first()
            if not p:
                raise ValueError(f"商品 {r.product_id} 不存在")
            p.current_price, p.updated_at = self._dec(r.suggested_price), datetime.utcnow()
            r.is_accepted, r.adopt_status = True, "ADOPTED"
            db.commit()
            return {"resultId": int(r.id), "productId": int(r.product_id), "suggestedPrice": _f(r.suggested_price)}
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

    def reject_result(self, result_id: int, reason: str = "") -> Dict[str, Any]:
        db: Session = SessionLocal()
        try:
            r = db.query(DecResult).filter(DecResult.id == result_id).first()
            if not r:
                raise ValueError(f"结果 {result_id} 不存在")
            r.is_accepted, r.adopt_status, r.reject_reason = False, "REJECTED", (reason or "").strip()[:500]
            db.commit()
            return {"resultId": int(r.id), "adoptStatus": r.adopt_status, "rejectReason": r.reject_reason or ""}
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

    def export_task_report_csv(self, task_id: int) -> bytes:
        rows = self.get_task_comparison(task_id)
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["taskId", "productId", "productTitle", "originalPrice", "suggestedPrice", "profitChange", "discountRate", "adoptStatus", "rejectReason"])
        for r in rows:
            w.writerow([r["taskId"], r["productId"], r["productTitle"], r["originalPrice"], r["suggestedPrice"], r["profitChange"], r["discountRate"], r["adoptStatus"], r.get("rejectReason") or ""])
        return out.getvalue().encode("utf-8-sig")


workflow_service = WorkflowService()


_orig_compose_data_agent_output = WorkflowService.compose_data_agent_output
_orig_compose_market_agent_output = WorkflowService.compose_market_agent_output
_orig_compose_risk_agent_output = WorkflowService.compose_risk_agent_output


def _compose_data_agent_output_with_price(self, name: str, result) -> str:
    text = _orig_compose_data_agent_output(self, name, result)
    display_price = getattr(result, "suggested_price", None)
    if str(getattr(result, "recommended_action", "")).lower() == "maintain":
        # Keep wording and numeric output consistent for maintain decisions.
        current_price = None
        try:
            current_price = float((result.analysis_details or {}).get("current_price"))
        except Exception:
            current_price = None
        if current_price is not None and current_price > 0:
            display_price = current_price
    if display_price is not None:
        text += f"\n4. 建议价格：{float(display_price):.2f}元"
    return text


def _compose_market_agent_output_with_price(self, name: str, result) -> str:
    text = _orig_compose_market_agent_output(self, name, result)
    if getattr(result, "suggested_price", None) is not None:
        text += f"\n4. 建议价格：{float(result.suggested_price):.2f}元"
    return text


def _compose_risk_agent_output_with_price(self, name: str, result) -> str:
    text = _orig_compose_risk_agent_output(self, name, result)
    if getattr(result, "suggested_price", None) is not None:
        text += f"\n4. 建议价格：{float(result.suggested_price):.2f}元"
    return text


WorkflowService.compose_data_agent_output = _compose_data_agent_output_with_price
WorkflowService.compose_market_agent_output = _compose_market_agent_output_with_price
WorkflowService.compose_risk_agent_output = _compose_risk_agent_output_with_price
